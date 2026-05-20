from __future__ import annotations

import argparse
import os
import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

try:
    import xlrd
except ModuleNotFoundError:
    vendor_path = str(Path(__file__).resolve().parent / "vendor")
    sys.path.insert(0, vendor_path)
    import xlrd
else:
    if not hasattr(xlrd, "open_workbook"):
        del sys.modules["xlrd"]
        vendor_path = str(Path(__file__).resolve().parent / "vendor")
        sys.path.insert(0, vendor_path)
        import xlrd


TARGET_COLUMNS = {
    "local": "L",
    "local 1": "L",
    "bari": "B",
    "bari 1": "B",
    "deposito": "D",
    "depósito": "D",
    "exporta": "E",
    "carrito": "C",
    "mayorista": "M",
    "control": "CO",
    "retail": "R",
}

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def norm_sku(value) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+", text):
        stripped = text.lstrip("0")
        return stripped or "0"
    return text


def norm_part(value, default: str = "") -> str:
    text = clean_text(value).upper()
    return text or default


def norm_qty(value) -> float:
    if value is None or clean_text(value) == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    text = clean_text(value).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0


def display_qty(value: float):
    return int(value) if float(value).is_integer() else value


def col_to_number(cell_ref: str) -> int:
    letters = re.match(r"[A-Z]+", cell_ref.upper()).group(0)
    number = 0
    for char in letters:
        number = number * 26 + ord(char) - 64
    return number


def number_to_col(number: int) -> str:
    letters = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def report_kind(path: Path) -> str | None:
    name = path.stem.lower()
    name = re.sub(r"[_\-]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    for token, column in TARGET_COLUMNS.items():
        if token in name:
            return column
    return None


def split_article(value) -> tuple[str, str]:
    text = clean_text(value).upper()
    if not text:
        return "", ""
    parts = text.split()
    sku = norm_sku(parts[0])
    color = norm_part(parts[1]) if len(parts) > 1 else ""
    return sku, color


def looks_like_header(row: Iterable) -> bool:
    values = [norm_part(v) for v in row]
    return "ARTÍCULO" in values or "ARTICULO" in values or "CANTIDAD" in values


def header_index(row: Iterable) -> dict[str, int]:
    out = {}
    for idx, value in enumerate(row):
        key = norm_part(value)
        if key in {"ARTÍCULO", "ARTICULO"} and "articulo" not in out:
            out["articulo"] = idx
        elif key == "TALLE" and "talle" not in out:
            out["talle"] = idx
        elif key == "CANTIDAD" and "cantidad" not in out:
            out["cantidad"] = idx
    return out


def iter_report_rows(path: Path):
    if path.suffix.lower() == ".xls":
        yield from iter_xls_report_rows(path)
    else:
        yield from iter_xlsx_report_rows(path)


def iter_xls_report_rows(path: Path):
    book = xlrd.open_workbook(path)
    for sheet in book.sheets():
        if sheet.nrows == 0:
            continue

        first_row = sheet.row_values(0)
        if looks_like_header(first_row):
            indexes = header_index(first_row)
            start_row = 1
        else:
            indexes = {"articulo": 0, "talle": 1, "cantidad": sheet.ncols - 1}
            start_row = 0

        if "articulo" not in indexes or "cantidad" not in indexes:
            continue

        for row_idx in range(start_row, sheet.nrows):
            row = sheet.row_values(row_idx)
            sku, color = split_article(row[indexes["articulo"]])
            talle_index = indexes.get("talle", -1)
            talle = norm_part(row[talle_index] if talle_index >= 0 else "", "ALL")
            qty = norm_qty(row[indexes["cantidad"]])
            if not sku or qty == 0:
                continue
            yield (sku, color, talle), qty


def iter_xlsx_report_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
        except StopIteration:
            continue

        if looks_like_header(first_row):
            indexes = header_index(first_row)
            data_rows = rows
        else:
            indexes = {"articulo": 0, "talle": 1, "cantidad": len(first_row) - 1}
            data_rows = [first_row, *rows]

        if "articulo" not in indexes or "cantidad" not in indexes:
            continue

        for row in data_rows:
            sku, color = split_article(row[indexes["articulo"]])
            talle_index = indexes.get("talle", -1)
            talle = norm_part(row[talle_index] if talle_index >= 0 else "", "ALL")
            qty = norm_qty(row[indexes["cantidad"]])
            if not sku or qty == 0:
                continue
            yield (sku, color, talle), qty


def collect_reports(report_paths: Iterable[Path]):
    totals = {column: defaultdict(float) for column in set(TARGET_COLUMNS.values())}
    stats = {}
    for path in report_paths:
        column = report_kind(path)
        if not column:
            stats[path.name] = {"column": None, "rows": 0}
            continue
        rows = 0
        for key, qty in iter_report_rows(path):
            totals[column][key] += qty
            rows += 1
        stats[path.name] = {"column": column, "rows": rows}
    return totals, stats


def stock_key(sheet, row: int, indexes: dict[str, int]) -> tuple[str, str, str]:
    return (
        norm_sku(sheet.cell(row, indexes["SKU"]).value),
        norm_part(sheet.cell(row, indexes["COLOR"]).value),
        norm_part(sheet.cell(row, indexes["TALLE"]).value, "ALL"),
    )


def read_shared_strings(zip_file: zipfile.ZipFile) -> list[str]:
    try:
        xml = zip_file.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(xml)
    strings = []
    for item in root.findall(f"{{{NS_MAIN}}}si"):
        parts = [node.text or "" for node in item.iter(f"{{{NS_MAIN}}}t")]
        strings.append("".join(parts))
    return strings


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.get("t")
    if cell_type == "s":
        value = cell.find(f"{{{NS_MAIN}}}v")
        if value is None or value.text is None:
            return ""
        index = int(float(value.text))
        return shared_strings[index] if 0 <= index < len(shared_strings) else ""
    if cell_type == "inlineStr":
        parts = [node.text or "" for node in cell.iter(f"{{{NS_MAIN}}}t")]
        return "".join(parts)
    value = cell.find(f"{{{NS_MAIN}}}v")
    return value.text if value is not None and value.text is not None else ""


def stock_sheet_path(zip_file: zipfile.ZipFile) -> str:
    workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rels = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))

    target_rel_id = None
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        if sheet.get("name") == "STOCK":
            target_rel_id = sheet.get(f"{{{NS_REL}}}id")
            break
    if not target_rel_id:
        raise ValueError("La planilla base no tiene una solapa llamada STOCK.")

    for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship"):
        if rel.get("Id") == target_rel_id:
            target = rel.get("Target", "")
            if target.startswith("/"):
                return target.lstrip("/")
            if target.startswith("worksheets/"):
                return f"xl/{target}"
            return f"xl/{target}"
    raise ValueError("No se pudo ubicar el XML interno de la solapa STOCK.")


def set_numeric_cell(row: ET.Element, row_number: int, column_number: int, value):
    cell_ref = f"{number_to_col(column_number)}{row_number}"
    cells = row.findall(f"{{{NS_MAIN}}}c")
    target = None
    insert_at = len(cells)
    for idx, cell in enumerate(cells):
        current_ref = cell.get("r", "")
        current_col = col_to_number(current_ref) if current_ref else idx + 1
        if current_col == column_number:
            target = cell
            break
        if current_col > column_number:
            insert_at = idx
            break

    if target is None:
        target = ET.Element(f"{{{NS_MAIN}}}c", {"r": cell_ref})
        row.insert(insert_at, target)

    target.attrib.pop("t", None)
    for child in list(target):
        target.remove(child)
    value_node = ET.SubElement(target, f"{{{NS_MAIN}}}v")
    value_node.text = str(value)


def update_stock_xlsx_fast(stock_path: Path, totals, output_path: Path):
    with zipfile.ZipFile(stock_path, "r") as source:
        shared_strings = read_shared_strings(source)
        sheet_path = stock_sheet_path(source)
        sheet_root = ET.fromstring(source.read(sheet_path))

        headers = {}
        matched = defaultdict(int)
        for row in sheet_root.findall(f".//{{{NS_MAIN}}}row"):
            row_number = int(row.get("r", "0"))
            values = {}
            for cell in row.findall(f"{{{NS_MAIN}}}c"):
                ref = cell.get("r", "")
                if not ref:
                    continue
                values[col_to_number(ref)] = cell_text(cell, shared_strings)

            if row_number == 1:
                headers = {norm_part(value): column for column, value in values.items()}
                required = ["SKU", "COLOR", "TALLE", *set(TARGET_COLUMNS.values())]
                missing = [name for name in required if name not in headers]
                if missing:
                    raise ValueError(f"Faltan columnas en STOCK: {', '.join(missing)}")
                continue

            key = (
                norm_sku(values.get(headers["SKU"], "")),
                norm_part(values.get(headers["COLOR"], "")),
                norm_part(values.get(headers["TALLE"], ""), "ALL"),
            )
            for column, data in totals.items():
                value = display_qty(data.get(key, 0))
                set_numeric_cell(row, row_number, headers[column], value)
                if key in data:
                    matched[column] += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        updated_sheet = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                if item.filename == sheet_path:
                    target.writestr(item, updated_sheet)
                else:
                    target.writestr(item, source.read(item.filename))

    return dict(matched)


def complete_stock(stock_path: Path, report_paths: Iterable[Path], output_path: Path):
    report_paths = [Path(p) for p in report_paths if Path(p).suffix.lower() in {".xls", ".xlsx"}]
    totals, stats = collect_reports(report_paths)
    matched = update_stock_xlsx_fast(stock_path, totals, output_path)

    return {
        "output": str(output_path),
        "reports": stats,
        "matched": matched,
        "loaded": {column: len(data) for column, data in totals.items()},
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Completa STOCK.xlsx con cantidades de reportes.")
    parser.add_argument("--stock", required=True)
    parser.add_argument("--reports-dir")
    parser.add_argument("--reports", nargs="*")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    report_paths = []
    if args.reports_dir:
        report_paths.extend(Path(args.reports_dir).glob("*.xls*"))
    if args.reports:
        report_paths.extend(Path(p) for p in args.reports)

    result = complete_stock(Path(args.stock), report_paths, Path(args.output))
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
